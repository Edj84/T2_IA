import os, operator
from openpyxl import Workbook
from openpyxl import load_workbook
from cogroo_interface import Cogroo

def abrirArquivo():        
        return load_workbook(filename = filepath)
        
def lerPerguntas():
                       
        for i in range(1, max_row+1):
                enunciado = planilha.cell(row = i, column = 2).value
                classe = planilha.cell(row = i, column = 4).value
              
                if enunciado != None and classe != None:
                        pergunta = [i, enunciado, classe]       
                        if(i%5 == 0):
                                perguntas_val.append(pergunta)
                        else:     
                                perguntas.append(pergunta)
                                if classe not in classes:
                                        classes.add(classe)                              

def lematizar(texto):
        return cogroo.lemmatize(texto)

def analisar(texto):
        return cogroo.analyze(texto)

def extrairTokens(texto):
        result = list()
        aux = texto.sentences[0].tokens
        for i in range (len(aux)):
                string = str(aux[i])
                split = string.split('#')
                if((("adv" in split[1]) or "v-inf" in split[1]) or "prop" in split[1]):
                        result.append(split[0])
        return result

def classificarPergunta(pergunta,isTrain):

        classe = pergunta[2]
        dicionario = dictDePerguntas_val
        if(isTrain):
                dicionario = dictDePerguntas
        if classe not in dicionario.keys():
                perguntasDaClasse = [pergunta]
                dicionario[classe] = perguntasDaClasse
        else:
                perguntasDaClasse = dicionario[classe]
                perguntasDaClasse.append(pergunta)

def criarBoW(dicionario):
        for classe in dicionario: #pego todas as perguntas de cada classe
                tokens = dict() #gero um dicionario dos tokens
                listaPerguntas = dicionario[classe]
                for i in range (len(listaPerguntas)):
                        pergunta = listaPerguntas[i]
                        pergunta = pergunta[4] #pego os tokens da pergunta atual
                        for token in pergunta:
                                if token not in tokens.keys():
                                        tokens[token] = 1
                                else:
                                        tokens[token] = tokens[token] + 1
                sorted_tokens = sorted(tokens.items(), key=operator.itemgetter(1),reverse=True)
                for j in range(5):
                        if(j < len(sorted_tokens)):
                                aux = sorted_tokens[j][0]
                                if(aux not in listaGeral):
                                        listaGeral.append(aux)
                        
def estruturar(listaPerguntas):
        for i in range(len(listaPerguntas)):
                vector = list()
                aux = listaPerguntas[i]
                for j in range(len(aux[4])):
                        palavra = aux[4][j]
                        for k in range(len(listaGeral)):
                                if(palavra == listaGeral[k]):
                                        vector.append(k)
                result = list()
                for w in range(len(listaGeral)):
                        result.append(0)
                for index in vector:
                        result[index] = 1
                listaPerguntas[i].append(result)

def criarOutput(lista,isTrain):
        path = outpath
        if(isTrain == False):
                path = os.getcwd()+"/weka_val.arff"
        f= open(path,"w+")
        f.write("@relation Arquivo\n\n")
        for i in range(len(listaGeral)):
                f.write("@attribute "+ listaGeral[i] +" numeric"+"\n")
        aux = str(classes)
        aux = aux.strip('[]')
        f.write("@attribute classe "+aux+"\n")
        f.write("\n@data\n")
        if(isTrain):
                for i in range(len(lista)):
                        aux = str(lista[i][5])
                        aux = aux.strip('[]')
                        f.write(aux+", "+lista[i][2] +"\n")
        else:
                for i in range(len(lista)):
                        aux = str(lista[i][5])
                        aux = aux.strip('[]')
                        f.write(aux+", ?\n")
        f.close()

def processarPalavras(lista,isTrain):
        for p in lista:
                p.append(lematizar(p[1]))
                morfo = analisar(p[3])
                tokens = extrairTokens(morfo)
                p.append(tokens)
                classificarPergunta(p,isTrain)

#variáveis globais
classes = set() #lista de todas as classes existentes
perguntas = list() #lista com todas as perguntas, suas infos e tokens relevantes
perguntas_val = list() #lista com todas as perguntas que foram separadas para validacao
dictDePerguntas = dict() #dicionario com as chaves sendo as classes, os dados as perguntas
dictDePerguntas_val = dict()
listaGeral = list() #lista geral dos termos que mais aparecem, BoW

filepath = os.getcwd()+"/corpus.xlsx"

outpath = os.getcwd()+"/weka_input.arff"
if(os.path.exists(outpath)):
        os.remove(outpath)

cogroo = Cogroo.Instance()

#Lendo arquivo
arquivo = abrirArquivo()
planilha = arquivo.active
max_row = planilha.max_row
max_column = planilha.max_column

#Lendo perguntas
lerPerguntas()

#Normalizando e separando as perguntas por classes
processarPalavras(perguntas,True)
processarPalavras(perguntas_val,False)

criarBoW(dictDePerguntas)

estruturar(perguntas)
estruturar(perguntas_val)

criarOutput(perguntas,True)
criarOutput(perguntas_val,False)

print("DONE")